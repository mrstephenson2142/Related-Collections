import openpyxl
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog
import os
import datetime


# Functions 

def convert_to_xml(csv_file):

    global root

    # create root element
    root = ET.Element("items")

    # add item elements with extref subelements


    for item in csv_file.iter_rows(min_row=2, values_only=True):
        
        if not any(item):
            continue
    

        v_title = str(item[0]).strip() if item[0] else None
        v_url = str(item[1]).strip() if item[1] else None
        item_elem = ET.SubElement(root, "item")
        if v_url == None:
            item_elem.text = v_title
        else:
            extref_elem = ET.SubElement(item_elem, "extref", {"xmlns:xlink": "http://www.w3.org/1999/xlink", "xlink:type": "simple", "xlink:show": "new", "xlink:actuate": "onRequest", "xlink:href": v_url})
            extref_elem.text = v_title


## Start 

# Vars
warnMsg = None

# Set the current directory as the starting location for the file picker
rootDir = tk.Tk()
rootDir.withdraw()
excel_file_path = filedialog.askopenfilename(initialdir=os.getcwd(), filetypes=[("Excel Files", "*.xlsx")])

# Load the workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Get the desired sheet (or the first sheet if "template" doesn't exist)
if "Template" in workbook.sheetnames:
    sheet = workbook["Template"]
else:
    sheet = workbook.active

# Create the XML document
#xml_doc = ET.Element("RelatedCollections")

# Convert Excel to XML 
convert_to_xml(sheet)

# Save the XML document
fileName = "related-collections-output"
filepath = os.getcwd()
file_suffix = datetime.datetime.now().strftime("%Y_%m_%d-%H%M_%S_%f")
fileName = fileName + "-" + file_suffix + ".txt"
fullpath = os.path.join(filepath, fileName)


# Format the output
output = ET.tostring(root, encoding="unicode")
output = output.replace("</item><item>", "</item>\n\n<item>")
output = output.replace("<items>", "")
output = output.replace("</items>", "")
#print(output.replace("</item><item>", "</item>\n\n<item>"))

# write output to file
with open(fullpath, "w") as outfile:
    outfile.write(output)

# Stop message
#print(f"Script completed. Results written to: {fullpath}", end="", flush=True) 
print(f"Script completed. Results written to: {fullpath}")



# Pause at the end if warnings happened during run. 
if warnMsg:
    print("Warnings occoured during run.")
    input("Press 'Enter' to exit and open the output file...")


# Open saved xml file remove the top and bottom two lines, then save it again.
# set the file name and open the file

# with open(fullpath, "r") as file:
#     # read the content of the file
#     content = file.readlines()

# remove the top two lines and bottom two lines of the file
# scontent = content[2:-1]

# save the modified content to the same file
# with open(fullpath, "w") as file:
#     file.writelines(content)

# Open the file 
os.startfile(fullpath)
